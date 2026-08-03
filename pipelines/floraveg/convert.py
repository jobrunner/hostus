#!/usr/bin/env python3
"""Life_form.xlsx (FloraVeg.EU) -> canonical name-list CSV.

Source file choice: FloraVeg.EU's download page (/download/) offers no
dedicated "taxon checklist" export — its per-topic Excel files are all trait
tables (life form, dispersal, disturbance indicator values, parasitism,
vegetation units, EUNIS habitats, ...) or the ESy expert system (explicitly
out of scope per the task brief). Of these, "Life_form.xlsx" is the
simplest and most complete per-taxon table: one row per taxon with a
SeqID and a set of Raunkiaer life-form flag columns, no vegetation/habitat
join needed. It is used here purely as a taxon/name list; the life-form
flags themselves are not part of this canonical output (a separate traits
pipeline, out of scope for this task, would be the place for them).

Sheet name observed: "FloraVegEU-2023-01-03" (doubles as the version/date
this list was last cut).
Columns observed: SeqID, FloraVeg.Taxon, <12 life-form flag columns>.
All 16402 rows are binomial-or-finer scientific names; no explicit
rank/status/accepted-taxon columns exist in this file (FloraVeg's own
checklist does not appear to record synonymy in this table) so:
  taxon          = FloraVeg.Taxon
  rank           = "" (not provided by this file; see summary word-count
                   histogram for a rough species/infraspecific split)
  status         = "accepted" (every row is a currently-used FloraVeg name;
                   no synonym flag exists in this table)
  accepted_taxon = "" (status is always accepted here)
  source_id      = SeqID
"""
import csv
import sys

import openpyxl


def main():
    in_path, out_path = sys.argv[1:3]

    wb = openpyxl.load_workbook(in_path, read_only=True, data_only=True)
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {name: i for i, name in enumerate(header) if name is not None}

    row_count = 0
    taxa = set()
    word_counts = {}

    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f, delimiter="|")
        w.writerow(["taxon", "rank", "status", "accepted_taxon", "source_id"])
        for r in rows:
            taxon = r[idx["FloraVeg.Taxon"]]
            if not taxon:
                continue
            seq_id = r[idx["SeqID"]]
            source_id = "" if seq_id is None else str(seq_id)
            w.writerow([taxon, "", "accepted", "", source_id])
            row_count += 1
            taxa.add(taxon)
            wc = len(str(taxon).split())
            word_counts[wc] = word_counts.get(wc, 0) + 1

    print(f"sheet={sheet_name}")
    print(f"rows={row_count} taxa={len(taxa)}")
    print("word_counts=" + ",".join(f"{k}:{v}" for k, v in sorted(word_counts.items())))


if __name__ == "__main__":
    main()
